from __future__ import annotations

import copy
import hashlib
import json
import logging
import threading
import time
import uuid
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Callable

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries


DEFAULT_SETTINGS = {
    "debounce_seconds": 5.0,
    "post_save_delay_seconds": 1.0,
    "read_retry_count": 3,
    "read_retry_delay_seconds": 1.2,
    "scan_interval_seconds": 2.0,
    "retry_locked_file_seconds": 5.0,
}


@dataclass
class ExcelSyncPaths:
    data_path: Path
    log_dir: Path
    template_data_path: Path | None = None

    @property
    def log_path(self) -> Path:
        return self.log_dir / "manager.log"


@dataclass
class SyncTask:
    id: str = field(default_factory=lambda: uuid.uuid4().hex)
    name: str = "新任务"
    enabled: bool = True
    source_file: str = ""
    source_sheet: str = ""
    source_mode: str = "whole_sheet"
    source_range: str = ""
    target_file: str = ""
    target_sheet: str = ""
    target_mode: str = "replace_sheet"
    target_start_cell: str = "A1"
    columns_by_header: list[str] = field(default_factory=list)
    header_row: int = 1
    data_start_row: int = 2
    copy_style: bool = True
    copy_column_widths: bool = True
    copy_row_heights: bool = True
    include_header: bool = True
    drop_empty_rows: bool = True
    formula_handling: str = "values"
    last_target_rows: int = 0
    last_target_cols: int = 0


@dataclass
class TaskRuntime:
    status: str = "idle"
    last_startsync_at: str = ""
    last_synced_at: str = ""
    last_error: str = ""
    last_change_at: str = ""
    pending_retry: bool = False


@dataclass
class SyncResult:
    path: Path
    rows_written: int
    cols_written: int


def create_demo_workbooks(sample_dir: Path) -> dict[str, Path]:
    sample_dir.mkdir(parents=True, exist_ok=True)
    source_path = sample_dir / "示例-源表.xlsx"
    target_path = sample_dir / "示例-目标表.xlsx"

    if not source_path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"
        rows = [
            ["订单号", "船名", "港口", "柜量", "状态", "备注"],
            ["SO-1001", "MORNING TIDE", "上海", 24, "待确认", "首票"],
            ["SO-1002", "PACIFIC LINK", "宁波", 18, "已确认", ""],
            ["SO-1003", "BLUE ORBIT", "青岛", 32, "待确认", "需复核"],
        ]
        for row in rows:
            ws.append(row)
        wb.save(source_path)
        wb.close()

    if not target_path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"
        ws["A1"] = "这里会写入同步结果"
        wb.save(target_path)
        wb.close()

    return {
        "source_path": source_path,
        "target_path": target_path,
    }


def list_sheets(source_file: str) -> list[str]:
    workbook = openpyxl.load_workbook(source_file, read_only=True, data_only=False)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def list_headers(
    source_file: str,
    sheet_name: str,
    header_row: int,
    source_mode: str = "whole_sheet",
    source_range: str = "",
) -> list[str]:
    workbook = openpyxl.load_workbook(source_file, read_only=True, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        ws = workbook[sheet_name]
        if source_mode == "custom_range":
            min_col, min_row, max_col, _ = range_boundaries(_normalize_excel_range(source_range))
            header_row = min_row
        else:
            min_col, max_col = 1, ws.max_column
        headers: list[str] = []
        for col_idx in range(min_col, max_col + 1):
            value = ws.cell(header_row, col_idx).value
            if value not in (None, ""):
                headers.append(str(value).strip())
        return headers
    finally:
        workbook.close()


def sync_task(task: SyncTask, settings: dict) -> SyncResult:
    if not task.source_file or not task.source_sheet:
        raise ValueError("必须先选择源 Excel 和工作表。")
    if task.source_mode == "custom_range" and not task.source_range:
        raise ValueError("当源区域模式为自定义区域时，必须填写区域，例如 F2:T13。")
    if not task.target_file or not task.target_sheet:
        raise ValueError("必须先选择目标 Excel 和目标工作表。")
    if task.target_mode == "write_from_cell" and not task.target_start_cell:
        raise ValueError("当目标写入方式为从单元格开始时，必须填写起始单元格。")
    if not task.columns_by_header:
        raise ValueError("至少要勾选一列需要同步的字段。")

    source_path = Path(task.source_file).resolve()
    if source_path.name.startswith("~$"):
        raise ValueError("请选择真实的 Excel 文件，不要选择 Office 临时锁文件。")
    target_path = Path(task.target_file).resolve()
    target_path.parent.mkdir(parents=True, exist_ok=True)

    post_save_delay = float(settings.get("post_save_delay_seconds", 0.0))
    if post_save_delay > 0:
        time.sleep(post_save_delay)

    retry_count = int(settings.get("read_retry_count", 3))
    retry_delay = float(settings.get("read_retry_delay_seconds", 1.2))
    last_error: Exception | None = None

    for _ in range(retry_count):
        wb = None
        wb_values = None
        out_wb = None
        try:
            wb = openpyxl.load_workbook(source_path, data_only=False)
            wb_values = openpyxl.load_workbook(source_path, data_only=True)
            if task.source_sheet not in wb.sheetnames:
                raise ValueError(f"源工作表不存在: {task.source_sheet}")

            src_ws = wb[task.source_sheet]
            src_ws_values = wb_values[task.source_sheet]
            source_columns = _get_selected_columns(src_ws, task)
            header_row, exported_rows = _get_export_row_indices(src_ws, task, source_columns)
            row_count = (1 if task.include_header else 0) + len(exported_rows)
            col_count = len(source_columns)

            out_wb, dst_ws, target_start_row, target_start_col = _prepare_target_workbook(task, row_count, col_count)
            target_row_idx = target_start_row
            target_row_map: dict[int, int] = {}

            if task.include_header:
                target_row_map[header_row] = target_row_idx
                for offset, source_col_idx in enumerate(source_columns):
                    src_cell = src_ws.cell(header_row, source_col_idx)
                    src_value_cell = src_ws_values.cell(header_row, source_col_idx)
                    dst_cell = dst_ws.cell(
                        target_row_idx,
                        target_start_col + offset,
                        _export_value(src_cell, src_value_cell, task),
                    )
                    if task.copy_style:
                        _copy_cell_style(src_cell, dst_cell)
                if task.copy_row_heights:
                    _apply_row_height(src_ws, dst_ws, header_row, target_row_idx)
                target_row_idx += 1

            for source_row_idx in exported_rows:
                target_row_map[source_row_idx] = target_row_idx
                for offset, source_col_idx in enumerate(source_columns):
                    src_cell = src_ws.cell(source_row_idx, source_col_idx)
                    src_value_cell = src_ws_values.cell(source_row_idx, source_col_idx)
                    dst_cell = dst_ws.cell(
                        target_row_idx,
                        target_start_col + offset,
                        _export_value(src_cell, src_value_cell, task),
                    )
                    if task.copy_style:
                        _copy_cell_style(src_cell, dst_cell)
                if task.copy_row_heights:
                    _apply_row_height(src_ws, dst_ws, source_row_idx, target_row_idx)
                target_row_idx += 1

            if task.copy_column_widths:
                _apply_column_widths(src_ws, dst_ws, source_columns, target_start_col)
            if task.copy_style:
                _clone_merged_cells(src_ws, dst_ws, source_columns, target_row_map, target_start_col)

            out_wb.save(target_path)
            return SyncResult(path=target_path, rows_written=row_count, cols_written=col_count)
        except PermissionError as exc:
            last_error = exc
            time.sleep(retry_delay)
        finally:
            if wb is not None:
                wb.close()
            if wb_values is not None:
                wb_values.close()
            if out_wb is not None:
                out_wb.close()

    if last_error is not None:
        raise last_error
    raise RuntimeError("同步失败。")


class SyncService:
    def __init__(self, paths: ExcelSyncPaths, status_callback: Callable[[], None] | None = None):
        self.paths = paths
        self.paths.log_dir.mkdir(parents=True, exist_ok=True)
        self.logger = self._build_logger()
        self.settings, tasks = self._load_data()
        self.tasks: list[SyncTask] = tasks
        self.runtime: dict[str, TaskRuntime] = {task.id: TaskRuntime() for task in tasks}
        self.status_callback = status_callback
        self._stop_event = threading.Event()
        self._thread: threading.Thread | None = None
        self._lock = threading.RLock()
        self._file_signatures: dict[str, tuple[int, int]] = {}
        self._sheet_signatures: dict[str, str] = {}
        self._pending_since: dict[str, float] = {}
        self._pending_retry_at: dict[str, float] = {}
        self._retry_sheet_signatures: dict[str, str] = {}

    def _build_logger(self) -> logging.Logger:
        logger_name = f"ams_excel_sync_{hash(self.paths.log_path)}"
        logger = logging.getLogger(logger_name)
        logger.setLevel(logging.INFO)
        logger.propagate = False
        if not logger.handlers:
            file_handler = logging.FileHandler(self.paths.log_path, encoding="utf-8")
            file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
            logger.addHandler(file_handler)
        return logger

    def _load_data(self) -> tuple[dict, list[SyncTask]]:
        if not self.paths.data_path.exists():
            self.paths.data_path.parent.mkdir(parents=True, exist_ok=True)
            if self.paths.template_data_path and self.paths.template_data_path.exists():
                self.paths.data_path.write_text(
                    self.paths.template_data_path.read_text(encoding="utf-8"),
                    encoding="utf-8",
                )
            else:
                self._save_data(DEFAULT_SETTINGS.copy(), [])
            return DEFAULT_SETTINGS.copy(), []

        raw = json.loads(self.paths.data_path.read_text(encoding="utf-8"))
        settings = DEFAULT_SETTINGS.copy()
        settings.update(raw.get("settings", {}))
        tasks = [SyncTask(**item) for item in raw.get("tasks", [])]
        return settings, tasks

    def _save_data(self, settings: dict, tasks: list[SyncTask]) -> None:
        payload = {
            "settings": settings,
            "tasks": [asdict(task) for task in tasks],
        }
        self.paths.data_path.parent.mkdir(parents=True, exist_ok=True)
        self.paths.data_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    def save(self) -> None:
        with self._lock:
            self._save_data(self.settings, self.tasks)

    def set_tasks(self, tasks: list[SyncTask]) -> None:
        with self._lock:
            existing_runtime = self.runtime
            old_tasks = {task.id: task for task in self.tasks}
            self.tasks = tasks
            self.runtime = {task.id: existing_runtime.get(task.id, TaskRuntime()) for task in tasks}
            keep_ids = {task.id for task in tasks}
            self._file_signatures = {task_id: value for task_id, value in self._file_signatures.items() if task_id in keep_ids}
            self._sheet_signatures = {task_id: value for task_id, value in self._sheet_signatures.items() if task_id in keep_ids}
            self._pending_since = {task_id: value for task_id, value in self._pending_since.items() if task_id in keep_ids}
            self._pending_retry_at = {task_id: value for task_id, value in self._pending_retry_at.items() if task_id in keep_ids}
            self._retry_sheet_signatures = {
                task_id: value for task_id, value in self._retry_sheet_signatures.items() if task_id in keep_ids
            }
            for task in tasks:
                previous = old_tasks.get(task.id)
                if previous is not None and asdict(previous) != asdict(task):
                    self._file_signatures.pop(task.id, None)
                    self._sheet_signatures.pop(task.id, None)
                    self._pending_since.pop(task.id, None)
                    self._pending_retry_at.pop(task.id, None)
                    self._retry_sheet_signatures.pop(task.id, None)
            self.save()
        self._notify()

    def start(self) -> None:
        if self._thread and self._thread.is_alive():
            return
        self._stop_event.clear()
        self._thread = threading.Thread(target=self._run_loop, daemon=True)
        self._thread.start()
        self.logger.info("Excel sync service started.")

    def stop(self) -> None:
        self._stop_event.set()
        if self._thread and self._thread.is_alive():
            self._thread.join(timeout=3)
        self.logger.info("Excel sync service stopped.")

    def is_running(self) -> bool:
        return bool(self._thread and self._thread.is_alive())

    def run_task_now(self, task_id: str) -> None:
        task = self.get_task(task_id)
        if task is None:
            return
        self._sync_one(task, manual=True)

    def get_task(self, task_id: str) -> SyncTask | None:
        with self._lock:
            for task in self.tasks:
                if task.id == task_id:
                    return task
        return None

    def list_runtime_rows(self) -> list[tuple[SyncTask, TaskRuntime]]:
        with self._lock:
            return [(task, self.runtime.get(task.id, TaskRuntime())) for task in self.tasks]

    def recent_log_lines(self, max_lines: int = 25) -> str:
        if not self.paths.log_path.exists():
            return ""
        lines = self.paths.log_path.read_text(encoding="utf-8", errors="ignore").splitlines()
        return "\n".join(lines[-max_lines:])

    def _notify(self) -> None:
        if self.status_callback is not None:
            self.status_callback()

    def _mark_status(self, task_id: str, **updates) -> None:
        with self._lock:
            runtime = self.runtime.setdefault(task_id, TaskRuntime())
            for key, value in updates.items():
                setattr(runtime, key, value)
        self._notify()

    def _run_loop(self) -> None:
        while not self._stop_event.is_set():
            self._scan_for_changes()
            self._process_pending()
            self._process_retries()
            self._stop_event.wait(float(self.settings.get("scan_interval_seconds", 2.0)))

    def _scan_for_changes(self) -> None:
        now = time.monotonic()
        with self._lock:
            tasks = list(self.tasks)
        for task in tasks:
            if not task.enabled or not task.source_file:
                continue
            source_path = Path(task.source_file)
            if not source_path.exists() or source_path.name.startswith("~$"):
                continue
            try:
                stat = source_path.stat()
            except OSError:
                continue
            signature = (stat.st_mtime_ns, stat.st_size)
            previous = self._file_signatures.get(task.id)
            self._file_signatures[task.id] = signature
            if previous is None:
                continue
            if signature != previous:
                self._pending_since[task.id] = now
                self._mark_status(
                    task.id,
                    status="检测到变化",
                    last_change_at=time.strftime("%Y-%m-%d %H:%M:%S"),
                )
                self.logger.info("Detected file change for task '%s'", task.name)

    def _process_pending(self) -> None:
        now = time.monotonic()
        debounce = float(self.settings.get("debounce_seconds", 5.0))
        due = [task_id for task_id, changed_at in self._pending_since.items() if now - changed_at >= debounce]
        for task_id in due:
            self._pending_since.pop(task_id, None)
            task = self.get_task(task_id)
            if task is None or not task.enabled:
                continue
            try:
                current_sheet_signature = _build_sheet_signature(task)
            except PermissionError:
                retry_delay = float(self.settings.get("retry_locked_file_seconds", 5.0))
                self._pending_retry_at[task.id] = time.monotonic() + retry_delay
                self._mark_status(
                    task.id,
                    status="等待文件解锁",
                    last_error="源文件或目标文件正被 Excel / WPS 占用。",
                    pending_retry=True,
                )
                continue
            except Exception as exc:
                self._mark_status(task.id, status="错误", last_error=str(exc), pending_retry=False)
                self.logger.exception("Task '%s' failed while checking sheet changes", task.name)
                continue

            previous_sheet_signature = self._sheet_signatures.get(task.id)
            if previous_sheet_signature == current_sheet_signature:
                self._mark_status(task.id, status="文件变化但同步区域未变", pending_retry=False, last_error="")
                self.logger.info("Task '%s' ignored file change because the selected sheet did not change.", task.name)
                continue
            self._sync_one(task, manual=False, source_signature=current_sheet_signature)

    def _process_retries(self) -> None:
        now = time.monotonic()
        due = [task_id for task_id, retry_at in self._pending_retry_at.items() if now >= retry_at]
        for task_id in due:
            self._pending_retry_at.pop(task_id, None)
            task = self.get_task(task_id)
            if task is None or not task.enabled:
                continue
            self._sync_one(task, manual=False, source_signature=self._retry_sheet_signatures.get(task_id))

    def _sync_one(self, task: SyncTask, manual: bool = False, source_signature: str | None = None) -> None:
        updates = {
            "status": "同步中",
            "pending_retry": False,
            "last_error": "",
        }
        if manual:
            updates["last_startsync_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
        self._mark_status(task.id, **updates)
        try:
            result = sync_task(task, self.settings)
            if source_signature is None:
                source_signature = _build_sheet_signature(task)
            self._sheet_signatures[task.id] = source_signature
            self._retry_sheet_signatures.pop(task.id, None)
            if task.last_target_rows != result.rows_written or task.last_target_cols != result.cols_written:
                task.last_target_rows = result.rows_written
                task.last_target_cols = result.cols_written
                self.save()
            try:
                synced_at = result.path.stat().st_mtime
                timestamp = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(synced_at))
            except OSError:
                timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
            self._mark_status(
                task.id,
                status=f"已同步 -> {result.path.name}",
                last_synced_at=timestamp,
                last_error="",
                pending_retry=False,
            )
            self.logger.info("Task '%s' synced to %s", task.name, result.path)
        except PermissionError:
            retry_delay = float(self.settings.get("retry_locked_file_seconds", 5.0))
            self._pending_retry_at[task.id] = time.monotonic() + retry_delay
            if source_signature is not None:
                self._retry_sheet_signatures[task.id] = source_signature
            self._mark_status(
                task.id,
                status="等待文件解锁",
                last_error="源文件或目标文件正被 Excel / WPS 占用。",
                pending_retry=True,
            )
            self.logger.warning("Task '%s' is waiting for file unlock.", task.name)
        except Exception as exc:
            self._mark_status(task.id, status="错误", last_error=str(exc), pending_retry=False)
            self.logger.exception("Task '%s' failed", task.name)


def _normalize_excel_range(cell_range: str) -> str:
    normalized = cell_range.replace("$", "").replace(" ", "").strip()
    if ":" not in normalized:
        raise ValueError("区域格式必须类似 F2:T13。")
    return normalized.upper()


def _normalize_cell(cell: str) -> str:
    normalized = cell.replace("$", "").replace(" ", "").strip()
    if not normalized:
        raise ValueError("单元格格式必须类似 B3。")
    return normalized.upper()


def _copy_cell_style(src_cell, dst_cell) -> None:
    dst_cell.font = copy.copy(src_cell.font)
    dst_cell.fill = copy.copy(src_cell.fill)
    dst_cell.border = copy.copy(src_cell.border)
    dst_cell.alignment = copy.copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy.copy(src_cell.protection)


def _row_is_empty(ws, row_idx: int, source_columns: list[int]) -> bool:
    for source_col in source_columns:
        if ws.cell(row_idx, source_col).value not in (None, ""):
            return False
    return True


def _apply_column_widths(src_ws, dst_ws, source_columns: list[int], dest_start_col: int) -> None:
    for offset, source_col_idx in enumerate(source_columns):
        source_letter = get_column_letter(source_col_idx)
        target_letter = get_column_letter(dest_start_col + offset)
        dst_ws.column_dimensions[target_letter].width = src_ws.column_dimensions[source_letter].width
        dst_ws.column_dimensions[target_letter].hidden = src_ws.column_dimensions[source_letter].hidden


def _apply_row_height(src_ws, dst_ws, source_row_idx: int, target_row_idx: int) -> None:
    src_dim = src_ws.row_dimensions[source_row_idx]
    if src_dim.height is not None:
        dst_ws.row_dimensions[target_row_idx].height = src_dim.height
    dst_ws.row_dimensions[target_row_idx].hidden = src_dim.hidden


def _clone_merged_cells(
    src_ws,
    dst_ws,
    source_columns: list[int],
    target_row_map: dict[int, int],
    dest_start_col: int,
) -> None:
    source_col_map = {source_col: dest_start_col + offset for offset, source_col in enumerate(source_columns)}
    for merged_range in src_ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if any(col not in source_col_map for col in range(min_col, max_col + 1)):
            continue
        if min_row not in target_row_map or max_row not in target_row_map:
            continue
        dst_ws.merge_cells(
            start_row=target_row_map[min_row],
            start_column=source_col_map[min_col],
            end_row=target_row_map[max_row],
            end_column=source_col_map[max_col],
        )


def _get_source_bounds(ws, task: SyncTask) -> tuple[int, int, int, int, int]:
    if task.source_mode == "custom_range":
        min_col, min_row, max_col, max_row = range_boundaries(_normalize_excel_range(task.source_range))
        header_row = min_row
        data_start_row = min_row + 1
        return min_col, max_col, max_row, header_row, data_start_row
    return 1, ws.max_column, ws.max_row, task.header_row, task.data_start_row


def _get_selected_columns(ws, task: SyncTask) -> list[int]:
    min_col, max_col, _, header_row, _ = _get_source_bounds(ws, task)
    header_map: dict[str, int] = {}
    for col_idx in range(min_col, max_col + 1):
        header_value = ws.cell(header_row, col_idx).value
        if header_value in (None, ""):
            continue
        header_map[str(header_value).strip()] = col_idx

    selected: list[int] = []
    missing: list[str] = []
    for header in task.columns_by_header:
        if header in header_map:
            selected.append(header_map[header])
        else:
            missing.append(header)
    if missing:
        raise ValueError(f"缺少这些表头: {', '.join(missing)}")
    return selected


def _get_export_row_indices(ws, task: SyncTask, source_columns: list[int]) -> tuple[int, list[int]]:
    _, _, max_row, header_row, data_start_row = _get_source_bounds(ws, task)
    exported_rows: list[int] = []
    for source_row_idx in range(data_start_row, max_row + 1):
        if task.drop_empty_rows and _row_is_empty(ws, source_row_idx, source_columns):
            continue
        exported_rows.append(source_row_idx)
    return header_row, exported_rows


def _export_value(src_formula_cell, src_value_cell, task: SyncTask):
    if task.formula_handling == "formulas":
        return src_formula_cell.value
    if src_formula_cell.data_type == "f":
        return src_value_cell.value
    return src_formula_cell.value


def _update_digest(digest, *parts) -> None:
    for part in parts:
        digest.update(str(part).encode("utf-8", errors="replace"))
        digest.update(b"\x1f")
    digest.update(b"\x1e")


def _build_sheet_signature(task: SyncTask) -> str:
    source_path = Path(task.source_file).resolve()
    workbook = None
    workbook_values = None
    try:
        workbook = openpyxl.load_workbook(source_path, data_only=False)
        workbook_values = openpyxl.load_workbook(source_path, data_only=True)
        if task.source_sheet not in workbook.sheetnames:
            raise ValueError(f"源工作表不存在: {task.source_sheet}")

        ws = workbook[task.source_sheet]
        ws_values = workbook_values[task.source_sheet]
        source_columns = _get_selected_columns(ws, task)
        header_row, exported_rows = _get_export_row_indices(ws, task, source_columns)
        digest = hashlib.sha256()
        _update_digest(
            digest,
            task.source_sheet,
            task.source_mode,
            task.source_range,
            task.header_row,
            task.data_start_row,
            task.include_header,
            task.drop_empty_rows,
            task.copy_style,
            task.copy_column_widths,
            task.copy_row_heights,
            task.formula_handling,
            "|".join(task.columns_by_header),
        )

        if task.include_header:
            for source_col_idx in source_columns:
                src_cell = ws.cell(header_row, source_col_idx)
                src_value_cell = ws_values.cell(header_row, source_col_idx)
                _update_digest(
                    digest,
                    "header",
                    header_row,
                    source_col_idx,
                    src_cell.data_type,
                    src_cell.value,
                    src_value_cell.value,
                    src_cell.number_format,
                    src_cell.style_id,
                )
            if task.copy_row_heights:
                row_dim = ws.row_dimensions[header_row]
                _update_digest(digest, "row-dim", header_row, row_dim.height, row_dim.hidden)

        if task.copy_column_widths:
            for source_col_idx in source_columns:
                col_letter = get_column_letter(source_col_idx)
                col_dim = ws.column_dimensions[col_letter]
                _update_digest(digest, "col-dim", source_col_idx, col_dim.width, col_dim.hidden)

        exported_row_set = set(exported_rows)
        for source_row_idx in exported_rows:
            if task.copy_row_heights:
                row_dim = ws.row_dimensions[source_row_idx]
                _update_digest(digest, "row-dim", source_row_idx, row_dim.height, row_dim.hidden)
            for source_col_idx in source_columns:
                src_cell = ws.cell(source_row_idx, source_col_idx)
                src_value_cell = ws_values.cell(source_row_idx, source_col_idx)
                _update_digest(
                    digest,
                    "cell",
                    source_row_idx,
                    source_col_idx,
                    src_cell.data_type,
                    src_cell.value,
                    src_value_cell.value,
                    src_cell.number_format,
                    src_cell.style_id,
                )

        if task.copy_style:
            source_column_set = set(source_columns)
            for merged_range in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                if any(col not in source_column_set for col in range(min_col, max_col + 1)):
                    continue
                row_set = set(range(min_row, max_row + 1))
                if task.include_header and header_row in row_set:
                    pass
                elif not row_set.issubset(exported_row_set):
                    continue
                _update_digest(digest, "merge", str(merged_range))

        return digest.hexdigest()
    finally:
        if workbook is not None:
            workbook.close()
        if workbook_values is not None:
            workbook_values.close()


def _clear_target_area(ws, start_row: int, start_col: int, rows_to_clear: int, cols_to_clear: int) -> None:
    if rows_to_clear <= 0 or cols_to_clear <= 0:
        return
    end_row = start_row + rows_to_clear - 1
    end_col = start_col + cols_to_clear - 1
    ranges_to_remove = []
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        overlaps = not (max_row < start_row or min_row > end_row or max_col < start_col or min_col > end_col)
        if overlaps:
            ranges_to_remove.append(str(merged_range))
    for merged_range in ranges_to_remove:
        ws.unmerge_cells(merged_range)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.value = None


def _prepare_target_workbook(task: SyncTask, rows_written: int, cols_written: int):
    target_path = Path(task.target_file).resolve()

    if task.target_mode == "write_from_cell":
        start_row, start_col = coordinate_to_tuple(_normalize_cell(task.target_start_cell or "A1"))
        if target_path.exists():
            workbook = openpyxl.load_workbook(target_path)
            if task.target_sheet in workbook.sheetnames:
                dst_ws = workbook[task.target_sheet]
            else:
                dst_ws = workbook.create_sheet(task.target_sheet)
        else:
            workbook = openpyxl.Workbook()
            dst_ws = workbook.active
            dst_ws.title = task.target_sheet
        rows_to_clear = max(task.last_target_rows, rows_written)
        cols_to_clear = max(task.last_target_cols, cols_written)
        _clear_target_area(dst_ws, start_row, start_col, rows_to_clear, cols_to_clear)
        return workbook, dst_ws, start_row, start_col

    if target_path.exists():
        workbook = openpyxl.load_workbook(target_path)
        if task.target_sheet in workbook.sheetnames:
            existing_sheet = workbook[task.target_sheet]
            sheet_index = workbook.sheetnames.index(task.target_sheet)
            workbook.remove(existing_sheet)
            dst_ws = workbook.create_sheet(task.target_sheet, sheet_index)
        else:
            dst_ws = workbook.create_sheet(task.target_sheet)
        return workbook, dst_ws, 1, 1

    workbook = openpyxl.Workbook()
    dst_ws = workbook.active
    dst_ws.title = task.target_sheet
    return workbook, dst_ws, 1, 1
