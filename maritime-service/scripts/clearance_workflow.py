#!/usr/bin/env python3
"""Customs-clearance workbook update workflow for the maritime-service skill."""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_ROOT = SCRIPT_DIR.parent
EXAMPLE_CASES_DIR = SKILL_ROOT / "examples" / "clearance_cases"
EXAMPLE_QUERY_DIR = SKILL_ROOT / "examples" / "clearance_query_results"
EXAMPLE_WORKBOOK_DIR = SKILL_ROOT / "examples" / "clearance_workbooks"
DATA_ROOT = Path(os.environ.get("AMS_DATA_ROOT", str(SKILL_ROOT / "output"))).resolve()
OUTPUT_ROOT = DATA_ROOT / "clearance"
UPDATED_DIR = OUTPUT_ROOT / "updated"
REPORT_DIR = OUTPUT_ROOT / "reports"

INFO_SHEET = "说明"
BUSINESS_SHEET = "业务表"
QUERY_SHEET = "查询结果"
SUMMARY_SHEET = "执行报告"

UPDATED_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
PENDING_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
MISSING_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")

BUSINESS_HEADERS = [
    "VOYAGE",
    "POL",
    "POD",
    "QTTY",
    "CARGO",
    "MARK",
    "备货",
    "B/L NO.",
    "PCS",
    "MT",
    "CBM",
    "FRT TON",
    "TERMS",
    "L/S/D",
    "打尺",
    "打尺联系人",
]

QUERY_HEADERS = [
    "B/L NO.",
    "网站通关状态",
    "是否已放行",
    "PCS",
    "MT",
    "CBM",
    "FRT TON",
    "查询时间",
    "原始备注",
]

REQUIRED_BUSINESS_COLUMNS = ["备货", "B/L NO.", "PCS", "MT"]
RELEASED_KEYWORDS = ("已放行", "放行", "已通关", "已结关", "released", "release", "cleared")
ALREADY_CLEARED_STATUSES = {"已通关", "已放行"}


class ClearanceError(Exception):
    """Base clearance workflow error."""


class ClearanceValidationError(ClearanceError):
    """Validation error with multiple user-facing messages."""

    def __init__(self, errors: list[str]) -> None:
        super().__init__("\n".join(errors))
        self.errors = errors


@dataclass
class ClearanceResult:
    workbook_path: Path
    report_json_path: Path
    report_txt_path: Path
    summary: dict[str, Any]


def compact_spaces(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_bl(value: Any) -> str:
    text = compact_spaces(value).upper()
    return re.sub(r"[\s\u3000]+", "", text)


def to_number_if_possible(value: Any) -> Any:
    text = compact_spaces(value)
    if not text:
        return ""
    try:
        number = float(text)
    except ValueError:
        return text
    if number.is_integer():
        return int(number)
    return number


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def make_output_name(source_name: str, suffix: str) -> str:
    stem = Path(source_name).stem
    return f"{stem}-{suffix}"


def set_header_style(sheet, headers: list[str]) -> None:
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.value = header
        cell.fill = HEADER_FILL


def create_workbook_template(
    output_path: Path,
    business_rows: list[dict[str, Any]] | None = None,
    query_rows: list[dict[str, Any]] | None = None,
) -> Path:
    workbook = Workbook()

    info_sheet = workbook.active
    info_sheet.title = INFO_SHEET
    info_sheet["A1"] = "这是 req2 自动查通关模板。真正需要填写的是“业务表”和“查询结果”两个工作表。"
    info_sheet["A2"] = "如果你还没有真网站自动查询能力，可以先把网站查到的结果手工粘贴到“查询结果”工作表。"
    info_sheet["A3"] = "脚本会只处理“备货”不是已通关且 B/L NO. 不为空的行。"
    info_sheet["A4"] = "如果“查询结果”里某票显示已放行，脚本会把“备货”改成已通关，并回填 PCS / MT。"
    info_sheet["A5"] = "建议先跑 docs/10-req2-如何验收.md 里的示例，再自己填真实表。"

    business_sheet = workbook.create_sheet(BUSINESS_SHEET)
    set_header_style(business_sheet, BUSINESS_HEADERS)
    for row in business_rows or []:
        business_sheet.append([row.get(header, "") for header in BUSINESS_HEADERS])

    query_sheet = workbook.create_sheet(QUERY_SHEET)
    set_header_style(query_sheet, QUERY_HEADERS)
    for row in query_rows or []:
        query_sheet.append([row.get(header, "") for header in QUERY_HEADERS])

    summary_sheet = workbook.create_sheet(SUMMARY_SHEET)
    summary_sheet["A1"] = "这里会在执行后自动生成摘要，不需要手工填写。"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def load_case_payload(path: Path) -> dict[str, Any]:
    payload = read_json(path)
    if not isinstance(payload.get("business_rows"), list):
        raise ClearanceValidationError([f"{path.name} 缺少 `business_rows`。"])
    return payload


def load_query_payload(path: Path) -> dict[str, Any]:
    payload = read_json(path)
    if not isinstance(payload.get("query_rows"), list):
        raise ClearanceValidationError([f"{path.name} 缺少 `query_rows`。"])
    return payload


def find_header_columns(sheet, required_headers: list[str]) -> dict[str, int]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {compact_spaces(value): index + 1 for index, value in enumerate(header_row) if compact_spaces(value)}
    missing = [header for header in required_headers if header not in header_map]
    if missing:
        raise ClearanceValidationError([f"工作表 `{sheet.title}` 缺少这些列：{missing}"])
    return header_map


def read_business_rows(sheet) -> list[dict[str, Any]]:
    headers = [compact_spaces(cell.value) for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(item not in (None, "") for item in row):
            continue
        payload = {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}
        payload["_row_index"] = row_index
        rows.append(payload)
    return rows


def read_query_rows(sheet) -> list[dict[str, Any]]:
    headers = [compact_spaces(cell.value) for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(item not in (None, "") for item in row):
            continue
        payload = {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}
        if normalize_bl(payload.get("B/L NO.", "")):
            rows.append(payload)
    return rows


def inject_query_rows(sheet, query_rows: list[dict[str, Any]]) -> None:
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for row in query_rows:
        sheet.append([row.get(header, "") for header in QUERY_HEADERS])


def is_released_row(row: dict[str, Any]) -> bool:
    release_flag = compact_spaces(row.get("是否已放行", "")).lower()
    if release_flag in {"y", "yes", "true", "1", "是", "已放行"}:
        return True
    status = compact_spaces(row.get("网站通关状态", "")).lower()
    return any(keyword in status for keyword in RELEASED_KEYWORDS)


def needs_lookup(status: Any) -> bool:
    normalized = compact_spaces(status)
    if not normalized:
        return True
    return normalized not in ALREADY_CLEARED_STATUSES


def build_query_index(query_rows: list[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], list[str]]:
    index: dict[str, dict[str, Any]] = {}
    duplicates: list[str] = []
    for row in query_rows:
        bl_no = normalize_bl(row.get("B/L NO.", ""))
        if not bl_no:
            continue
        if bl_no in index:
            duplicates.append(bl_no)
            continue
        index[bl_no] = row
    return index, duplicates


def apply_fill(cell, fill: PatternFill) -> None:
    cell.fill = copy(fill)


def update_summary_sheet(sheet, summary: dict[str, Any]) -> None:
    if sheet.max_row > 1:
        sheet.delete_rows(1, sheet.max_row)
    rows = [
        ("执行时间", summary["run_at"]),
        ("输入工作簿", summary["source_workbook"]),
        ("总业务行数", summary["total_rows"]),
        ("待查询行数", summary["candidate_rows"]),
        ("已更新为已通关", summary["updated_rows"]),
        ("原本已通关跳过", summary["already_cleared_rows"]),
        ("仍未放行", summary["still_pending_rows"]),
        ("找不到查询结果", summary["missing_query_rows"]),
        ("重复查询结果", ", ".join(summary["duplicate_query_bls"]) or "无"),
    ]
    for row_index, (label, value) in enumerate(rows, start=1):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)

    sheet.cell(row=len(rows) + 2, column=1, value="已更新提单")
    for offset, item in enumerate(summary["updated_items"], start=len(rows) + 3):
        sheet.cell(row=offset, column=1, value=item["bl_no"])
        sheet.cell(row=offset, column=2, value=item["new_status"])
        sheet.cell(row=offset, column=3, value=item["pcs"])
        sheet.cell(row=offset, column=4, value=item["mt"])

    pending_start = len(rows) + max(len(summary["updated_items"]), 1) + 5
    sheet.cell(row=pending_start, column=1, value="仍未放行或未命中")
    for offset, item in enumerate(summary["pending_or_missing_items"], start=pending_start + 1):
        sheet.cell(row=offset, column=1, value=item["bl_no"])
        sheet.cell(row=offset, column=2, value=item["reason"])


def render_report_text(summary: dict[str, Any]) -> str:
    lines = [
        "req2 自动查通关执行报告",
        f"执行时间: {summary['run_at']}",
        f"输入工作簿: {summary['source_workbook']}",
        "",
        "统计摘要",
        f"- 总业务行数: {summary['total_rows']}",
        f"- 待查询行数: {summary['candidate_rows']}",
        f"- 已更新为已通关: {summary['updated_rows']}",
        f"- 原本已通关跳过: {summary['already_cleared_rows']}",
        f"- 仍未放行: {summary['still_pending_rows']}",
        f"- 找不到查询结果: {summary['missing_query_rows']}",
        "",
        "已更新项目",
    ]
    if summary["updated_items"]:
        for item in summary["updated_items"]:
            lines.append(f"- {item['bl_no']} -> {item['new_status']} | PCS={item['pcs']} | MT={item['mt']}")
    else:
        lines.append("- 无")

    lines.append("")
    lines.append("未更新项目")
    if summary["pending_or_missing_items"]:
        for item in summary["pending_or_missing_items"]:
            lines.append(f"- {item['bl_no']} -> {item['reason']}")
    else:
        lines.append("- 无")
    return "\n".join(lines)


def update_clearance_workbook(
    workbook_path: Path,
    output_dir: Path,
    query_rows_override: list[dict[str, Any]] | None = None,
) -> ClearanceResult:
    workbook_path = workbook_path.resolve()
    output_dir = output_dir.resolve()
    workbook = load_workbook(workbook_path)
    if BUSINESS_SHEET not in workbook.sheetnames or QUERY_SHEET not in workbook.sheetnames:
        raise ClearanceValidationError([f"工作簿必须包含 `{BUSINESS_SHEET}` 和 `{QUERY_SHEET}` 两个工作表。"])

    business_sheet = workbook[BUSINESS_SHEET]
    query_sheet = workbook[QUERY_SHEET]
    summary_sheet = workbook[SUMMARY_SHEET] if SUMMARY_SHEET in workbook.sheetnames else workbook.create_sheet(SUMMARY_SHEET)

    business_columns = find_header_columns(business_sheet, REQUIRED_BUSINESS_COLUMNS)
    find_header_columns(query_sheet, ["B/L NO.", "网站通关状态", "PCS", "MT"])

    if query_rows_override is not None:
        inject_query_rows(query_sheet, query_rows_override)

    business_rows = read_business_rows(business_sheet)
    query_rows = read_query_rows(query_sheet)
    query_index, duplicate_query_bls = build_query_index(query_rows)

    updated_items: list[dict[str, Any]] = []
    pending_or_missing_items: list[dict[str, Any]] = []
    candidate_rows = 0
    already_cleared_rows = 0
    still_pending_rows = 0
    missing_query_rows = 0

    for row in business_rows:
        row_index = row["_row_index"]
        bl_no_raw = row.get("B/L NO.", "")
        normalized_bl = normalize_bl(bl_no_raw)
        if not normalized_bl:
            continue

        current_status = compact_spaces(row.get("备货", ""))
        if not needs_lookup(current_status):
            already_cleared_rows += 1
            continue

        candidate_rows += 1
        matched = query_index.get(normalized_bl)
        status_cell = business_sheet.cell(row=row_index, column=business_columns["备货"])
        pcs_cell = business_sheet.cell(row=row_index, column=business_columns["PCS"])
        mt_cell = business_sheet.cell(row=row_index, column=business_columns["MT"])

        if not matched:
            missing_query_rows += 1
            apply_fill(status_cell, MISSING_FILL)
            pending_or_missing_items.append({"bl_no": compact_spaces(bl_no_raw), "reason": "查询结果中没有这票提单"})
            continue

        site_status = compact_spaces(matched.get("网站通关状态", ""))
        if is_released_row(matched):
            status_cell.value = "已通关"
            apply_fill(status_cell, UPDATED_FILL)

            pcs_value = to_number_if_possible(matched.get("PCS", ""))
            mt_value = to_number_if_possible(matched.get("MT", ""))
            if pcs_value != "":
                pcs_cell.value = pcs_value
                apply_fill(pcs_cell, UPDATED_FILL)
            if mt_value != "":
                mt_cell.value = mt_value
                apply_fill(mt_cell, UPDATED_FILL)

            updated_items.append(
                {
                    "bl_no": compact_spaces(bl_no_raw),
                    "new_status": "已通关",
                    "pcs": pcs_value,
                    "mt": mt_value,
                    "site_status": site_status,
                }
            )
        else:
            still_pending_rows += 1
            apply_fill(status_cell, PENDING_FILL)
            pending_or_missing_items.append(
                {
                    "bl_no": compact_spaces(bl_no_raw),
                    "reason": f"网站状态仍未放行：{site_status or '空'}",
                }
            )

    source_name = workbook_path.name
    output_name = make_output_name(source_name, "updated.xlsx")
    output_dir.mkdir(parents=True, exist_ok=True)
    UPDATED_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    summary = {
        "run_at": now_iso(),
        "source_workbook": source_name,
        "total_rows": len(business_rows),
        "candidate_rows": candidate_rows,
        "updated_rows": len(updated_items),
        "already_cleared_rows": already_cleared_rows,
        "still_pending_rows": still_pending_rows,
        "missing_query_rows": missing_query_rows,
        "duplicate_query_bls": duplicate_query_bls,
        "updated_items": updated_items,
        "pending_or_missing_items": pending_or_missing_items,
    }

    update_summary_sheet(summary_sheet, summary)
    workbook_out = (output_dir / output_name).resolve()
    workbook.save(workbook_out)

    report_json_path = (REPORT_DIR / make_output_name(source_name, "report.json")).resolve()
    report_txt_path = (REPORT_DIR / make_output_name(source_name, "report.txt")).resolve()
    write_json(report_json_path, summary)
    report_txt_path.write_text(render_report_text(summary), encoding="utf-8")

    return ClearanceResult(workbook_out, report_json_path, report_txt_path, summary)


def build_example_workbooks() -> list[Path]:
    built: list[Path] = []
    EXAMPLE_WORKBOOK_DIR.mkdir(parents=True, exist_ok=True)

    for case_path in sorted(EXAMPLE_CASES_DIR.glob("*.json")):
        case_payload = load_case_payload(case_path)
        query_name = compact_spaces(case_payload.get("query_file", ""))
        query_rows: list[dict[str, Any]] = []
        if query_name:
            query_payload = load_query_payload(EXAMPLE_QUERY_DIR / query_name)
            query_rows = query_payload["query_rows"]
        output_path = EXAMPLE_WORKBOOK_DIR / f"{case_path.stem}.xlsx"
        built.append(create_workbook_template(output_path, case_payload["business_rows"], query_rows))

    blank_path = EXAMPLE_WORKBOOK_DIR / "blank-clearance-template.xlsx"
    built.append(create_workbook_template(blank_path, None, None))
    return built


def build_examples() -> list[ClearanceResult]:
    results: list[ClearanceResult] = []
    build_example_workbooks()
    for case_path in sorted(EXAMPLE_CASES_DIR.glob("*.json")):
        case_payload = load_case_payload(case_path)
        query_payload = load_query_payload(EXAMPLE_QUERY_DIR / case_payload["query_file"])
        workbook_path = EXAMPLE_WORKBOOK_DIR / f"{case_path.stem}.xlsx"
        results.append(update_clearance_workbook(workbook_path, UPDATED_DIR, query_payload["query_rows"]))
    return results


def verify_examples() -> list[str]:
    verification_dir = OUTPUT_ROOT / "_verification"
    if verification_dir.exists():
        shutil.rmtree(verification_dir)
    verification_dir.mkdir(parents=True, exist_ok=True)

    messages: list[str] = []
    for case_path in sorted(EXAMPLE_CASES_DIR.glob("*.json")):
        case_payload = load_case_payload(case_path)
        query_payload = load_query_payload(EXAMPLE_QUERY_DIR / case_payload["query_file"])
        workbook_path = EXAMPLE_WORKBOOK_DIR / f"{case_path.stem}.xlsx"
        if not workbook_path.exists():
            create_workbook_template(workbook_path, case_payload["business_rows"], query_payload["query_rows"])
        result = update_clearance_workbook(workbook_path, verification_dir, query_payload["query_rows"])
        summary = result.summary

        expected = case_payload.get("expected_summary", {})
        for key, value in expected.items():
            if summary.get(key) != value:
                raise ClearanceError(f"{case_path.name} 校验失败：`{key}` 期望 {value!r}，实际 {summary.get(key)!r}")

        updated_bl_set = {item["bl_no"] for item in summary["updated_items"]}
        for bl_no in case_payload.get("expected_updated_bls", []):
            if bl_no not in updated_bl_set:
                raise ClearanceError(f"{case_path.name} 校验失败：应更新的提单 `{bl_no}` 未更新。")

        pending_bl_set = {item["bl_no"] for item in summary["pending_or_missing_items"]}
        for bl_no in case_payload.get("expected_pending_bls", []):
            if bl_no not in pending_bl_set:
                raise ClearanceError(f"{case_path.name} 校验失败：应保留未更新的提单 `{bl_no}` 不在报告中。")

        messages.append(f"{case_path.name} -> 通过")
    return messages


def load_query_override(path: Path | None) -> list[dict[str, Any]] | None:
    if path is None:
        return None
    payload = load_query_payload(path)
    return payload["query_rows"]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    parser_workbook = subparsers.add_parser("from-workbook", help="从 req2 工作簿更新通关状态")
    parser_workbook.add_argument("--input", required=True, help="req2 Excel 工作簿路径")
    parser_workbook.add_argument("--query-file", help="可选：覆盖工作簿中的查询结果 JSON")
    parser_workbook.add_argument("--output-dir", default=str(UPDATED_DIR), help="输出目录")

    parser_template = subparsers.add_parser("make-workbook-template", help="生成 req2 空白模板")
    parser_template.add_argument("--output", required=True, help="输出的 xlsx 路径")
    parser_template.add_argument("--example-case", help="可选：使用某个示例 case 预填业务表和查询结果")

    subparsers.add_parser("build-examples", help="生成 req2 示例工作簿并输出示例更新结果")
    subparsers.add_parser("verify-examples", help="校验 req2 示例结果")
    return parser


def print_result(result: ClearanceResult) -> None:
    print(f"[OK] 更新后的工作簿: {result.workbook_path}")
    print(f"[OK] JSON 报告: {result.report_json_path}")
    print(f"[OK] 文本报告: {result.report_txt_path}")
    print(
        "[INFO] 统计: 待查询 {candidate_rows} | 已更新 {updated_rows} | 未放行 {still_pending_rows} | 未命中 {missing_query_rows}".format(
            **result.summary
        )
    )


def main(argv: list[str] | None = None) -> int:
    sys.stdout.reconfigure(encoding="utf-8")
    parser = build_parser()
    args = parser.parse_args(argv)

    try:
        if args.command == "make-workbook-template":
            business_rows = None
            query_rows = None
            if args.example_case:
                case_payload = load_case_payload(Path(args.example_case))
                business_rows = case_payload["business_rows"]
                query_rows = load_query_payload(EXAMPLE_QUERY_DIR / case_payload["query_file"])["query_rows"]
            output_path = create_workbook_template(Path(args.output), business_rows, query_rows)
            print(f"[OK] req2 Excel 模板已生成: {output_path}")
            return 0

        if args.command == "build-examples":
            results = build_examples()
            print("[OK] req2 示例结果已全部生成。")
            for result in results:
                print_result(result)
            return 0

        if args.command == "verify-examples":
            messages = verify_examples()
            print("[OK] req2 示例校验通过：")
            for message in messages:
                print(f" - {message}")
            return 0

        if args.command == "from-workbook":
            result = update_clearance_workbook(
                workbook_path=Path(args.input),
                output_dir=Path(args.output_dir),
                query_rows_override=load_query_override(Path(args.query_file)) if args.query_file else None,
            )
            print_result(result)
            return 0

        raise ClearanceError(f"不支持的命令：{args.command}")
    except ClearanceValidationError as exc:
        print("[ERROR] 输入校验失败：")
        for error in exc.errors:
            print(f" - {error}")
        return 1
    except ClearanceError as exc:
        print(f"[ERROR] {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
