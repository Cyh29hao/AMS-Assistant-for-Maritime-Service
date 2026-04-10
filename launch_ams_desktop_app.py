from __future__ import annotations

import argparse
import json
import queue
import sys
import threading
import traceback
from pathlib import Path
from tkinter import messagebox

import ttkbootstrap as ttk
from ttkbootstrap.constants import BOTH, LEFT, X

THIS_DIR = Path(__file__).resolve().parent
ROOT = THIS_DIR
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from desktop_app.runtime import APP_NAME, AmsOperations, AppConfig, help_assets_dir, release_assets_dir, skill_root


class UpdateInstallerWindow:
    def __init__(self, manifest_path: Path) -> None:
        self.manifest_path = manifest_path
        self.ops = AmsOperations(AppConfig.default())
        self.queue: queue.Queue[tuple[str, object]] = queue.Queue()
        self.running = True
        self.debug_log_path = self.manifest_path.parent / "updater-debug.log"
        self.write_debug(f"START manifest={self.manifest_path}")

        self.window = ttk.Window(themename="flatly")
        self.window.title(f"{APP_NAME} Updater")
        self.window.geometry("560x280")
        self.window.minsize(560, 280)
        self.window.maxsize(700, 420)
        self.window.protocol("WM_DELETE_WINDOW", self.on_close_requested)

        self.progress_var = ttk.DoubleVar(value=0)
        self.percent_var = ttk.StringVar(value="0%")
        self.status_var = ttk.StringVar(value="正在准备更新…")
        self.detail_var = ttk.StringVar(value=str(self.manifest_path.parent))

        self.build_ui()
        self.window.after(120, self.process_queue)
        self.start_worker()

    def build_ui(self) -> None:
        outer = ttk.Frame(self.window, padding=(22, 20, 22, 20))
        outer.pack(fill=BOTH, expand=True)

        card = ttk.Frame(outer, padding=22, bootstyle="light", borderwidth=1, relief="solid")
        card.pack(fill=BOTH, expand=True)

        ttk.Label(
            card,
            text="AUTO UPDATE",
            bootstyle="inverse-primary",
            font=("Microsoft YaHei UI", 9, "bold"),
            padding=(10, 5),
        ).pack(anchor="w")
        ttk.Label(card, text=f"{APP_NAME} 正在更新", font=("Microsoft YaHei UI", 18, "bold"), bootstyle="primary").pack(anchor="w", pady=(12, 4))
        ttk.Label(
            card,
            text="这一步会等待旧版本退出、安装新版本，并自动重新打开应用。你的工作区、设置和登录态不会丢失。",
            bootstyle="secondary",
            wraplength=500,
            justify="left",
        ).pack(anchor="w")

        progress_row = ttk.Frame(card)
        progress_row.pack(fill=X, pady=(18, 8))
        ttk.Progressbar(
            progress_row,
            variable=self.progress_var,
            maximum=100,
            bootstyle="success-striped",
        ).pack(side=LEFT, fill=X, expand=True)
        ttk.Label(progress_row, textvariable=self.percent_var, font=("Microsoft YaHei UI", 10, "bold"), bootstyle="primary").pack(side=LEFT, padx=(12, 0))

        ttk.Label(card, textvariable=self.status_var, font=("Microsoft YaHei UI", 12, "bold")).pack(anchor="w", pady=(10, 4))
        ttk.Label(card, textvariable=self.detail_var, bootstyle="secondary", wraplength=500, justify="left").pack(anchor="w")

        tips = ttk.Frame(card)
        tips.pack(fill=X, pady=(18, 0))
        ttk.Label(
            tips,
            text="请不要在这一步手动删除程序文件，也不用重复点击检查更新。",
            bootstyle="secondary",
            wraplength=500,
            justify="left",
        ).pack(anchor="w")

    def start_worker(self) -> None:
        def worker() -> None:
            try:
                result = self.ops.apply_prepared_update(
                    self.manifest_path,
                    progress_callback=lambda payload: self.queue.put(("progress", payload)),
                )
            except Exception:
                self.queue.put(("error", traceback.format_exc()))
                return
            self.queue.put(("success", result))

        threading.Thread(target=worker, daemon=True).start()

    def process_queue(self) -> None:
        try:
            while True:
                kind, payload = self.queue.get_nowait()
                if kind == "progress":
                    progress = payload if isinstance(payload, dict) else {}
                    value = float(progress.get("value", 0))
                    self.progress_var.set(value)
                    self.percent_var.set(f"{value:.0f}%")
                    self.status_var.set(str(progress.get("message", "正在更新…")))
                    self.detail_var.set(str(progress.get("detail", "")))
                    self.write_debug(
                        "PROGRESS "
                        f"value={value:.1f} message={self.status_var.get()} detail={self.detail_var.get()}"
                    )
                elif kind == "success":
                    result = payload if isinstance(payload, dict) else {}
                    self.running = False
                    self.progress_var.set(100)
                    self.percent_var.set("100%")
                    self.status_var.set("更新完成，正在打开新版本。")
                    self.detail_var.set(str(result.get("restarted_executable", "")))
                    self.write_debug(f"SUCCESS result={result}")
                    self.window.after(1400, self.window.destroy)
                elif kind == "error":
                    self.running = False
                    error_text = str(payload)
                    self.status_var.set("更新失败")
                    self.detail_var.set("可以重试一次检查更新；如果仍失败，请保留当前文件夹并反馈日志。")
                    self.write_debug(f"ERROR {error_text}")
                    messagebox.showerror("自动更新失败", error_text)
        except queue.Empty:
            pass
        finally:
            if self.window.winfo_exists():
                self.window.after(120, self.process_queue)

    def on_close_requested(self) -> None:
        if self.running:
            self.write_debug("CLOSE_BLOCKED running=true")
            messagebox.showinfo("正在更新", "自动更新正在进行，请稍等片刻。")
            return
        self.write_debug("CLOSE_ALLOWED")
        self.window.destroy()

    def run(self) -> int:
        self.window.mainloop()
        self.write_debug("EXIT")
        return 0

    def write_debug(self, message: str) -> None:
        self.debug_log_path.parent.mkdir(parents=True, exist_ok=True)
        with self.debug_log_path.open("a", encoding="utf-8") as fh:
            fh.write(message + "\n")


def run_self_test(output_path: Path, workspace_root: Path | None = None) -> int:
    import openpyxl
    import shutil

    from desktop_app.excel_sync_engine import DEFAULT_SETTINGS as DEFAULT_SYNC_SETTINGS
    from desktop_app.excel_sync_engine import SyncTask, create_demo_workbooks, sync_task
    from desktop_app.private_pack import build_private_pack

    config = AppConfig.default() if workspace_root is None else AppConfig(workspace_root=str(workspace_root.resolve()))
    ops = AmsOperations(config)
    info = ops.ensure_workspace()

    example_dir = skill_root() / "examples" / "workbooks"
    example_workbook = next(example_dir.glob("domestic-forwarder-*.xlsx"))
    req1_result = ops.contract_generate_from_file(example_workbook)

    sync_demo = create_demo_workbooks(ops.sync_examples_dir, force=True)
    sync_result = sync_task(
        SyncTask(
            name="自检任务",
            source_file=str(sync_demo["source_path"]),
            source_sheet="Orders",
            target_file=str(sync_demo["target_path"]),
            target_sheet="Export",
            column_selection_mode="exclude",
            columns_by_header=["备注"],
            header_row=1,
            data_start_row=2,
        ),
        DEFAULT_SYNC_SETTINGS,
    )

    sync_workbook = openpyxl.load_workbook(sync_result.path)
    try:
        sync_ws = sync_workbook["Export"]
        hidden_row_preserved = sync_ws.row_dimensions[3].hidden
        blank_row_preserved = sync_ws.max_row >= 5
        excluded_column_removed = sync_ws.max_column == 5
    finally:
        sync_workbook.close()

    private_test_root = ops.workspace_root / "_self_test_private_pack"
    if private_test_root.exists():
        shutil.rmtree(private_test_root)
    private_source = private_test_root / "source"
    private_dist = private_test_root / "dist"
    (private_source / "maritime-service" / "assets" / "contract_templates").mkdir(parents=True, exist_ok=True)
    (private_source / "maritime-service" / "scripts").mkdir(parents=True, exist_ok=True)
    (private_source / "desktop_app" / "release_assets" / "help").mkdir(parents=True, exist_ok=True)
    (private_source / "private-pack.json").write_text(
        json.dumps(
            {
                "pack_id": "ams-self-test-pack",
                "display_name": "AMS Self Test Pack",
                "version": "self-test",
                "description": "Built automatically by desktop self-test.",
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    public_template = next((skill_root() / "assets" / "contract_templates").glob("*.docx"))
    copied_template = private_source / "maritime-service" / "assets" / "contract_templates" / public_template.name
    shutil.copy2(public_template, copied_template)
    (private_source / "maritime-service" / "scripts" / "contract_template_registry.json").write_text(
        json.dumps(
            {
                "domestic_forwarder": {
                    "display_name": "Self Test Private Template",
                    "template_file": public_template.name,
                    "contract_no_pattern": "SELFTEST-{yyyymm}-{vessel_name_upper}",
                }
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    (private_source / "maritime-service" / "scripts" / "clearance_site_config.json").write_text(
        json.dumps(
            {
                "site_name": "Self Test Private Site",
                "base_url": "https://example.invalid/self-test",
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    (private_source / "desktop_app" / "release_assets" / "help" / "private-pack.html").write_text(
        "<!doctype html><html lang='zh-CN'><body><h1>Private help override</h1></body></html>",
        encoding="utf-8",
    )
    private_pack_path = private_dist / "ams-self-test.amspack"
    private_dist.mkdir(parents=True, exist_ok=True)
    private_summary = build_private_pack(private_source, private_pack_path, "self-test-password")
    private_pack_demo_created = private_pack_path.exists()
    private_install = ops.install_private_pack(private_pack_path, "self-test-password")
    private_status = ops.private_pack_status()
    private_help_path = ops.resolve_help_page_path("private")
    private_help_override_exists = private_help_path.exists()
    private_contract_module = ops.load_contract_module()
    private_site_module = ops.load_site_module()
    private_contract_pattern = private_contract_module.load_registry()["domestic_forwarder"]["contract_no_pattern"]
    private_site_name = private_site_module.load_config()["site_name"]
    private_clear = ops.clear_private_pack()
    private_restored_public = not private_clear.get("installed")
    private_help_fallback_exists = ops.resolve_help_page_path("private").exists()
    shutil.rmtree(private_test_root, ignore_errors=True)

    report = {
        "success": True,
        "frozen": bool(getattr(sys, "frozen", False)),
        "workspace_root": info["workspace_root"],
        "contract_input_path": info["contract_input_path"],
        "clearance_input_path": info["clearance_input_path"],
        "req1_document_path": req1_result["document_path"],
        "req1_summary_path": req1_result["summary_path"],
        "req1_latest_document_path": req1_result["latest_document_path"],
        "req1_latest_summary_path": req1_result["latest_summary_path"],
        "req1_latest_document_exists": Path(req1_result["latest_document_path"]).exists(),
        "req1_latest_summary_exists": Path(req1_result["latest_summary_path"]).exists(),
        "sync_demo_source_path": str(sync_demo["source_path"]),
        "sync_demo_target_path": str(sync_demo["target_path"]),
        "sync_target_exists": sync_result.path.exists(),
        "sync_hidden_row_preserved": hidden_row_preserved,
        "sync_blank_row_preserved": blank_row_preserved,
        "sync_excluded_column_removed": excluded_column_removed,
        "help_index_exists": (help_assets_dir() / "index.html").exists(),
        "private_pack_example_source_exists": ops.private_pack_example_source_dir.exists(),
        "private_pack_feature_available": True,
        "private_pack_demo_created": private_pack_demo_created,
        "private_pack_demo_is_temporary": True,
        "private_pack_demo_path": str(private_pack_path),
        "private_pack_demo_name": private_summary.display_name,
        "private_pack_demo_features": private_summary.features,
        "private_pack_installed": private_status["installed"],
        "private_pack_mode_label": private_status["mode_label"],
        "private_pack_contract_override_active": private_status["contract_registry_active"],
        "private_pack_clearance_override_active": private_status["clearance_site_config_active"],
        "private_pack_help_override_active": private_status["help_overrides_active"],
        "private_pack_contract_pattern": private_contract_pattern,
        "private_pack_site_name": private_site_name,
        "private_pack_help_path": str(private_help_path),
        "private_pack_help_override_exists": private_help_override_exists,
        "private_pack_install_result_name": private_install.get("display_name") or private_install.get("summary", {}).get("display_name", ""),
        "private_pack_cleared_back_to_public": private_restored_public,
        "private_pack_help_fallback_exists": private_help_fallback_exists,
        "guide_exists": (release_assets_dir() / "应用使用说明.html").exists(),
        "updater_exists": (Path(sys.executable).resolve().parent / "AMS-Assistant-Updater.exe").exists()
        if getattr(sys, "frozen", False)
        else False,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8-sig")
    return 0


def run_gui() -> int:
    from desktop_app.app import main as gui_main

    gui_main()
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--self-test-output")
    parser.add_argument("--self-test-workspace")
    parser.add_argument("--apply-update-manifest")
    return parser


def write_update_bootstrap_log(manifest_arg: str, stage: str) -> None:
    try:
        manifest_path = Path(manifest_arg).resolve()
        log_path = manifest_path.parent / "updater-bootstrap.log"
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with log_path.open("a", encoding="utf-8") as fh:
            fh.write(f"{stage} argv={sys.argv}\n")
    except Exception:
        pass


def main() -> int:
    parser = build_parser()
    args, _unknown = parser.parse_known_args()
    if args.self_test:
        output_path = Path(args.self_test_output or "ams-desktop-self-test.json").resolve()
        workspace_root = Path(args.self_test_workspace).resolve() if args.self_test_workspace else None
        return run_self_test(output_path, workspace_root)
    if args.apply_update_manifest:
        write_update_bootstrap_log(args.apply_update_manifest, "ENTER_UPDATE_MODE")
        return UpdateInstallerWindow(Path(args.apply_update_manifest).resolve()).run()
    return run_gui()


if __name__ == "__main__":
    raise SystemExit(main())
